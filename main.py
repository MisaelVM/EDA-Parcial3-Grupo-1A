from openpyxl import load_workbook
from mtree import MTree

wb = load_workbook(filename = 'covid_kaggle_dataset.xlsx')
sheet = wb.active

rows = sheet.max_row
print(rows)

class Patient:
	def __init__(self, id, data, dimensions):
		self.id = id
		self.data = []
		for i in data:
			if i == None:
				self.data.append(float(0))
			elif i == "negative":
				self.data.append(float(0))
			elif i == "positive":
				self.data.append(float(1))
			else:
				self.data.append(float(i))
		self.dimensions = dimensions
	
	def __repr__(self):
		return "-------------------------------------------------------\nPatient ID %s\nSARS-Cov-2 test result %s\nPatient age quantile %s\nHematocrit %s\nPlatelets %s\nMean platelet volume %s\nMCHC %s\nLeukocytes %s\nBasophils %s\nEosinophils %s\nMonocytes %s\nProteina C reativa mg/dL %s\n" % (
			self.id, self.data[0], self.data[1], self.data[2], self.data[3], self.data[4], self.data[5], self.data[6], self.data[7] ,self.data[8], self.data[9], self.data[10]
		)

def manhattan_d(a, b):
	if len(a.dimensions) < len(b.dimensions):
		dimensions = a.dimensions
	else:
		dimensions = b.dimensions

	distance = 0
	for i in dimensions:
		distance += abs(a.data[i] - b.data[i])

	return distance

def main():
	tree = MTree(manhattan_d, max_node_size=12)
	patients = []

	for i in range(2, rows + 1):
		id = sheet['A' + str(i)].value			#Patient ID
		data = []
		data.append(sheet['C' + str(i)].value)	#SARS-Cov-2 test result		0
		data.append(sheet['B' + str(i)].value)	#Patient age quantile		1
		data.append(sheet['G' + str(i)].value)	#Hematocrit					2
		data.append(sheet['I' + str(i)].value)	#Platelets					3
		data.append(sheet['J' + str(i)].value)	#Mean platelet volume		4
		data.append(sheet['M' + str(i)].value)	#MCHC						5
		data.append(sheet['N' + str(i)].value)	#Leukocytes					6
		data.append(sheet['O' + str(i)].value)	#Basophils					7
		data.append(sheet['Q' + str(i)].value)	#Eosinophils				8
		data.append(sheet['S' + str(i)].value)	#Monocytes					9
		data.append(sheet['AP' + str(i)].value)	#Proteina C reativa mg/dL	10
		dimensions = [ 0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10 ]
		patients.append(Patient(id, data, dimensions))

	tree.add_all(patients)

	while True:
		print("Ingrese las dimensiones:")
		print("0) SARS-Cov-2 test result")
		print("1) Patient age quantile")
		print("2) Hematocrit")
		print("3) Platelets")
		print("4) Mean platelet volume")
		print("5) MCHC")
		print("6) Leukocytes")
		print("7) Basophils")
		print("8) Eosinophils")
		print("9) Monocytes")
		print("10) Proteina C reativa mg/dL")
		dims = [int(i) for i in input("-1) Salir\n").split()]

		if dims[0] == -1:
			break

		print("Ingrese los respectivos datos:")
		id = "query"
		data = [float(0)] * 11
		for i in dims:
			data[i] = float(input())
		radius = float(input("Ingrese el radio de búsqueda:\n"))

		query_obj = Patient(id, data, dims)

		query = tree.search_in_range(query_obj, radius)
		for obj in query:
			print(obj)
		print(len(query), 'objects found')


if __name__ == '__main__':
	main()
